"""
Application Inventaire Biodiversité — Version 3.0
===================================================
Université André Salifou de Zinder
Développeur : Moctar Maman Laouan Nouhou
Email       : Moctar3490@gmail.com
Année       : 2025 — 2026

Fonctionnalités :
- Formulaire complet en 4 onglets (Localisation, Taxonomie, Observation, Preuves)
- GPS automatique (latitude, longitude, altitude, précision)
- Date/heure automatique
- Multi-photos + sélection galerie + caméra
- Enregistrement audio (chemin fichier)
- Listes déroulantes (milieu, sexe, stade, comportement, protocole...)
- Statut de validation
- Export Excel complet
- Synchronisation Google Sheets
- Graphiques (4 charts)
- Recherche/filtre dans la liste
- Modification et suppression
- 100% offline — sync quand réseau disponible
- Migration automatique de l'ancienne base de données
"""

import os
import sqlite3
import requests
from datetime import datetime
import math

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.image import Image
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.spinner import Spinner
from kivy.graphics import Color, Rectangle, Ellipse, Line
from kivy.metrics import dp

try:
    from plyer import gps, camera
    PLYER_OK = True
except Exception:
    PLYER_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

# ═══════════════════════════════════════════
# LISTES DÉROULANTES STANDARDISÉES
# ═══════════════════════════════════════════

MILIEUX = [
    "Sélectionner…",
    "Forêt dense", "Forêt claire / Savane boisée", "Savane arbustive",
    "Savane herbacée", "Prairie / Steppe", "Zone humide / Mare",
    "Fleuve / Rivière", "Galerie forestière", "Zone agricole / Champ",
    "Zone urbaine / Péri-urbaine", "Montagne / Plateau rocheux",
    "Désert / Zone aride", "Autre",
]

REGNES = [
    "Sélectionner…",
    "Mammifère", "Oiseau", "Reptile", "Amphibien", "Poisson",
    "Insecte", "Arachnide", "Mollusque", "Autre Invertébré",
    "Plante / Flore", "Champignon", "Autre",
]

STADES_VIE = [
    "Sélectionner…",
    "Adulte", "Juvénile", "Larve", "Poussin / Jeune",
    "Oeuf", "Graine / Fruit", "Indéterminé",
]

SEXES = [
    "Sélectionner…",
    "Mâle", "Femelle", "Les deux", "Indéterminé",
]

TYPES_INDICE = [
    "Sélectionner…",
    "Observation visuelle directe", "Chant / Cri", "Empreinte",
    "Fèces / Déjections", "Terrier / Nid / Gîte", "Cadavre",
    "Reste de repas / Pelote", "Piège photographique", "Autre",
]

ABONDANCES = [
    "Sélectionner…",
    "1 individu", "2-5 individus", "6-10 individus",
    "11-50 individus", "50-100 individus", "> 100 individus",
    "Nombre exact (voir notes)",
]

COMPORTEMENTS = [
    "Sélectionner…",
    "Repos / Inactif", "Alimentation / Chasse", "Nidification / Reproduction",
    "Transit / Déplacement", "Chant / Parade nuptiale", "Territorial",
    "Fuite / Alarme", "Autre",
]

PROTOCOLES = [
    "Sélectionner…",
    "Inventaire opportuniste", "Inventaire flash", "Point d'écoute",
    "Transect linéaire", "Piégeage photographique", "Piégeage entomologique",
    "Pêche électrique", "Relevé botanique", "Autre",
]

STATUTS_VALIDATION = ["En attente", "Validé", "Rejeté", "À vérifier"]

# ═══════════════════════════════════════════
# COULEURS
# ═══════════════════════════════════════════

VERT       = (0.13, 0.55, 0.13, 1)
VERT2      = (0.18, 0.70, 0.18, 1)
BLANC      = (1,    1,    1,    1)
ROUGE      = (0.8,  0.15, 0.15, 1)
ORANGE     = (0.9,  0.5,  0.0,  1)
BLEU       = (0.20, 0.60, 0.86, 1)
VIOLET     = (0.56, 0.27, 0.68, 1)
INDIGO     = (0.2,  0.2,  0.5,  1)
GRIS_FOND  = (0.97, 0.97, 0.97, 1)
ONGLET_COLORS = [VERT, BLEU, ORANGE, VIOLET]
CHART_COLORS  = [
    (0.13,0.55,0.13,1),(0.20,0.60,0.86,1),(0.90,0.49,0.13,1),
    (0.76,0.15,0.22,1),(0.56,0.27,0.68,1),(0.17,0.63,0.60,1),
    (0.93,0.80,0.07,1),(0.61,0.35,0.71,1),(0.33,0.42,0.18,1),
]

# ═══════════════════════════════════════════
# GPS
# ═══════════════════════════════════════════

gps_data = {"lat": "", "lon": "", "alt": "", "precision": ""}
gps_label_ref = None

def gps_callback(**kwargs):
    gps_data["lat"]       = str(round(kwargs.get("lat",       0), 6))
    gps_data["lon"]       = str(round(kwargs.get("lon",       0), 6))
    gps_data["alt"]       = str(round(kwargs.get("altitude",  0), 1))
    gps_data["precision"] = str(round(kwargs.get("accuracy",  0), 1))
    if gps_label_ref:
        gps_label_ref.text = (
            f"GPS  Lat:{gps_data['lat']}  Lon:{gps_data['lon']}  "
            f"Alt:{gps_data['alt']}m  Prec:{gps_data['precision']}m"
        )

def start_gps():
    if not PLYER_OK:
        return
    try:
        gps.configure(on_location=gps_callback)
        gps.start(minTime=1000, minDistance=0)
    except Exception as e:
        print(f"GPS : {e}")

# ═══════════════════════════════════════════
# BASE DE DONNÉES
# ═══════════════════════════════════════════

conn   = None
cursor = None

NOUVELLES_COLONNES = [
    ("altitude",           "TEXT"),
    ("precision_gps",      "TEXT"),
    ("type_milieu",        "TEXT"),
    ("stade_vie",          "TEXT"),
    ("sexe",               "TEXT"),
    ("date_obs",           "TEXT"),
    ("heure_obs",          "TEXT"),
    ("type_indice",        "TEXT"),
    ("abondance",          "TEXT"),
    ("comportement",       "TEXT"),
    ("nb_individus",       "TEXT"),
    ("audio",              "TEXT"),
    ("observateur",        "TEXT"),
    ("protocole",          "TEXT"),
    ("statut_validation",  "TEXT"),
    ("notes",              "TEXT"),
]

def init_db():
    global conn, cursor
    try:
        from android.storage import app_storage_path  # type: ignore
        db_dir = app_storage_path()
    except Exception:
        db_dir = os.path.dirname(os.path.abspath(__file__))

    db_path = os.path.join(db_dir, "inventaire.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS species (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            nom_scientifique TEXT,
            nom_local        TEXT,
            type             TEXT,
            lieu             TEXT,
            latitude         TEXT,
            longitude        TEXT,
            description      TEXT,
            photos           TEXT
        )
    """)
    conn.commit()

    # Migration : photo -> photos
    try:
        cols = [r[1] for r in cursor.execute("PRAGMA table_info(species)")]
        if "photo" in cols and "photos" not in cols:
            cursor.execute("ALTER TABLE species RENAME COLUMN photo TO photos")
            conn.commit()
    except Exception:
        pass

    # Migration V3 : nouvelles colonnes
    try:
        cols = [r[1] for r in cursor.execute("PRAGMA table_info(species)")]
        for col_name, col_type in NOUVELLES_COLONNES:
            if col_name not in cols:
                cursor.execute(f"ALTER TABLE species ADD COLUMN {col_name} {col_type}")
        conn.commit()
    except Exception as e:
        print(f"Migration V3 : {e}")

    return db_path

# ═══════════════════════════════════════════
# UTILITAIRES UI
# ═══════════════════════════════════════════

def styled_btn(text, color=None, height=None, text_color=BLANC):
    if color is None: color = VERT
    if height is None: height = dp(46)
    return Button(
        text=text, size_hint_y=None, height=height,
        background_normal="", background_color=color,
        color=text_color, bold=True, font_size=dp(14),
    )

def make_header(title, color=None):
    if color is None: color = VERT
    header = BoxLayout(size_hint_y=None, height=dp(56), padding=[dp(16), dp(8)])
    with header.canvas.before:
        Color(*color)
        r = Rectangle(pos=header.pos, size=header.size)
    header.bind(pos=lambda i,v: setattr(r,'pos',v),
                size=lambda i,v: setattr(r,'size',v))
    header.add_widget(Label(text=title, bold=True, font_size=dp(18), color=BLANC))
    return header

def show_popup(title, message, color=None):
    if color is None: color = VERT
    content = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10))
    lbl = Label(text=message, text_size=(dp(280), None),
                halign="center", color=(0.9,0.9,0.9,1))
    lbl.bind(texture_size=lbl.setter("size"))
    btn = styled_btn("OK", color=color)
    content.add_widget(lbl)
    content.add_widget(btn)
    popup = Popup(title=title, content=content,
                  size_hint=(0.88, None), height=dp(250),
                  title_color=BLANC, background="",
                  background_color=(0.1,0.1,0.1,0.96))
    btn.bind(on_press=popup.dismiss)
    popup.open()

def make_spinner(values, default=None):
    if default is None: default = values[0]
    return Spinner(
        text=default, values=values,
        size_hint_y=None, height=dp(44),
        background_normal="", background_color=(0.88,0.88,0.88,1),
        color=(0.1,0.1,0.1,1), font_size=dp(13),
    )

def section_lbl(text):
    l = Label(text=text, bold=True, font_size=dp(13), color=VERT,
              size_hint_y=None, height=dp(30), halign="left")
    l.bind(size=lambda i,v: setattr(l,'text_size',v))
    return l

def field_lbl(text):
    l = Label(text=text, font_size=dp(12), color=(0.45,0.45,0.45,1),
              size_hint_y=None, height=dp(20), halign="left")
    l.bind(size=lambda i,v: setattr(l,'text_size',v))
    return l

def txt(hint, multiline=False, height=dp(44)):
    return TextInput(hint_text=hint, multiline=multiline,
                     size_hint_y=None, height=height, font_size=dp(13))

# ═══════════════════════════════════════════
# EXPORT EXCEL
# ═══════════════════════════════════════════

def export_excel():
    if not OPENPYXL_OK:
        show_popup("Erreur", "openpyxl non installé.", ROUGE)
        return
    try:
        from android.storage import primary_external_storage_path  # type: ignore
        export_dir = os.path.join(primary_external_storage_path(), "Download")
    except Exception:
        export_dir = os.path.dirname(os.path.abspath(__file__))

    os.makedirs(export_dir, exist_ok=True)
    path = os.path.join(export_dir, "inventaire_biodiversite.xlsx")
    cursor.execute("SELECT * FROM species")
    rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"

    headers = [
        "ID","Nom scientifique","Nom local","Règne/Groupe","Lieu",
        "Latitude","Longitude","Altitude (m)","Précision GPS (m)",
        "Type de milieu","Description","Stade de vie","Sexe",
        "Date observation","Heure observation","Type d'indice",
        "Abondance","Nb individus","Comportement",
        "Observateur","Protocole","Statut validation",
        "Photos","Audio","Notes",
    ]
    hfill = PatternFill("solid", fgColor="1C8C1C")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(h)+4)

    def g(r, i, d=""):
        return r[i] if len(r) > i and r[i] else d

    for row in rows:
        ws.append([
            g(row,0), g(row,1), g(row,2), g(row,3), g(row,4),
            g(row,5), g(row,6), g(row,9), g(row,10), g(row,11),
            g(row,7), g(row,12), g(row,13), g(row,14), g(row,15),
            g(row,16), g(row,17), g(row,19), g(row,18), g(row,20),
            g(row,21), g(row,22), g(row,8), g(row,23), g(row,24),
        ])

    wb.save(path)
    show_popup("Export réussi", f"Fichier :\n{path}", VERT)

# ═══════════════════════════════════════════
# SYNC GOOGLE SHEETS
# ═══════════════════════════════════════════

GOOGLE_SHEET_URL = "COLLE_ICI_TON_URL_GOOGLE_APPS_SCRIPT"

def sync_google_sheets():
    if GOOGLE_SHEET_URL == "COLLE_ICI_TON_URL_GOOGLE_APPS_SCRIPT":
        show_popup("Configuration requise",
                   "Configure l'URL dans\nParametres Sync.", ORANGE)
        return
    cursor.execute("SELECT * FROM species")
    rows = cursor.fetchall()
    if not rows:
        show_popup("Sync", "Aucune espèce à envoyer.", ORANGE)
        return

    def g(r, i, d=""):
        return r[i] if len(r) > i and r[i] else d

    success = errors = 0
    for row in rows:
        payload = {
            "nom_scientifique":  g(row,1),
            "nom_local":         g(row,2),
            "type":              g(row,3),
            "lieu":              g(row,4),
            "latitude":          g(row,5),
            "longitude":         g(row,6),
            "altitude":          g(row,9),
            "precision_gps":     g(row,10),
            "type_milieu":       g(row,11),
            "description":       g(row,7),
            "stade_vie":         g(row,12),
            "sexe":              g(row,13),
            "date_obs":          g(row,14),
            "heure_obs":         g(row,15),
            "type_indice":       g(row,16),
            "abondance":         g(row,17),
            "comportement":      g(row,18),
            "nb_individus":      g(row,19),
            "audio":             g(row,20),
            "observateur":       g(row,21),
            "protocole":         g(row,22),
            "statut_validation": g(row,23),
            "notes":             g(row,24),
            "nb_photos": len([p for p in g(row,8).split("|") if p.strip()]),
        }
        try:
            r = requests.post(GOOGLE_SHEET_URL, json=payload,
                              timeout=15, allow_redirects=True)
            if r.status_code == 200:
                success += 1
            else:
                errors += 1
        except Exception as e:
            print(f"Sync : {e}")
            errors += 1

    if errors == 0:
        show_popup("Sync réussi", f"{success} observation(s) envoyée(s) !", VERT)
    else:
        show_popup("Sync partiel",
                   f"{success} OK  /  {errors} erreur(s)\n"
                   "Vérifie ta connexion.", ORANGE)

# ═══════════════════════════════════════════
# GRAPHIQUES
# ═══════════════════════════════════════════

class PieChartWidget(BoxLayout):
    def __init__(self, data_dict, title="", **kwargs):
        super().__init__(orientation="vertical", **kwargs)
        self.data = data_dict
        self.title = title
        self.bind(size=self._draw, pos=self._draw)

    def _draw(self, *args):
        self.canvas.before.clear()
        if not self.data: return
        total = sum(self.data.values()) or 1
        cx = self.x + self.width/2
        cy = self.y + self.height/2 - dp(20)
        r  = min(self.width, self.height) * 0.32
        start = 0
        with self.canvas.before:
            for i,(lbl,val) in enumerate(self.data.items()):
                Color(*CHART_COLORS[i%len(CHART_COLORS)])
                Ellipse(pos=(cx-r,cy-r), size=(r*2,r*2),
                        angle_start=start, angle_end=start+(val/total)*360)
                start += (val/total)*360
            Color(1,1,1,1)
            ir = r*0.48
            Ellipse(pos=(cx-ir,cy-ir), size=(ir*2,ir*2))
        self.clear_widgets()
        self.add_widget(Label(text=self.title, bold=True,
                              size_hint_y=None, height=dp(28), color=(0.1,0.1,0.1,1)))
        self.add_widget(BoxLayout(size_hint_y=0.55))
        leg = GridLayout(cols=2, size_hint_y=None, height=dp(22*len(self.data)))
        for i,(lbl,val) in enumerate(self.data.items()):
            leg.add_widget(Label(text="■", color=CHART_COLORS[i%len(CHART_COLORS)],
                                  size_hint_x=None, width=dp(22), font_size=dp(16)))
            leg.add_widget(Label(text=f"{lbl} ({val})", font_size=dp(11),
                                  color=(0.1,0.1,0.1,1), halign="left"))
        self.add_widget(leg)


class BarChartWidget(BoxLayout):
    def __init__(self, data_dict, title="", **kwargs):
        super().__init__(orientation="vertical", **kwargs)
        self.data = data_dict
        self.title = title
        self.bind(size=self._draw, pos=self._draw)

    def _draw(self, *args):
        self.canvas.before.clear()
        if not self.data: return
        items  = list(self.data.items())
        max_v  = max(v for _,v in items) or 1
        n      = len(items)
        mg     = dp(40)
        aw     = self.width - mg*2
        ah     = self.height - mg*2 - dp(40)
        bw     = aw/n*0.6
        gap    = aw/n
        with self.canvas.before:
            Color(*GRIS_FOND)
            Rectangle(pos=self.pos, size=self.size)
            Color(0.3,0.3,0.3,1)
            Line(points=[self.x+mg,self.y+mg,self.x+mg,self.y+mg+ah], width=1.5)
            Line(points=[self.x+mg,self.y+mg,self.x+mg+aw,self.y+mg], width=1.5)
            for i,(lbl,val) in enumerate(items):
                bh = (val/max_v)*ah
                Color(*CHART_COLORS[i%len(CHART_COLORS)])
                Rectangle(pos=(self.x+mg+i*gap+gap*0.2, self.y+mg), size=(bw,bh))
        self.clear_widgets()
        self.add_widget(Label(text=self.title, bold=True,
                              size_hint_y=None, height=dp(28), color=(0.1,0.1,0.1,1)))
        self.add_widget(BoxLayout(size_hint_y=0.72))
        row = GridLayout(cols=n, size_hint_y=None, height=dp(24),
                         padding=[dp(40),0,dp(10),0])
        for lbl,val in items:
            row.add_widget(Label(text=f"{lbl}\n({val})", font_size=dp(9),
                                  color=(0.1,0.1,0.1,1), halign="center"))
        self.add_widget(row)

# ═══════════════════════════════════════════
# ÉCRAN : MENU
# ═══════════════════════════════════════════

class MenuScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        root = BoxLayout(orientation="vertical")

        # Header
        header = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(120),
                           padding=[dp(12),dp(8)], spacing=dp(2))
        with header.canvas.before:
            Color(*VERT)
            hr = Rectangle(pos=header.pos, size=header.size)
        header.bind(pos=lambda i,v: setattr(hr,'pos',v),
                    size=lambda i,v: setattr(hr,'size',v))
        for text, fs, col, h in [
            ("Université André Salifou de Zinder", dp(11), (0.85,1,0.85,1), dp(20)),
            ("Inventaire Biodiversité",             dp(22), BLANC,           dp(36)),
            ("Moctar Maman Laouan Nouhou",          dp(11), (0.85,1,0.85,1), dp(18)),
            ("Moctar3490@gmail.com",                dp(10), (0.75,1,0.75,1), dp(16)),
        ]:
            header.add_widget(Label(text=text, font_size=fs, color=col,
                                     bold=(fs==dp(22)), size_hint_y=None, height=h))
        root.add_widget(header)

        # GPS
        self.gps_lbl = Label(text="GPS : recherche en cours…",
                              size_hint_y=None, height=dp(24),
                              color=(0.35,0.35,0.35,1), font_size=dp(11))
        global gps_label_ref
        gps_label_ref = self.gps_lbl
        root.add_widget(self.gps_lbl)

        # Boutons
        scroll = ScrollView()
        btns = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(16),
                          size_hint_y=None)
        btns.bind(minimum_height=btns.setter("height"))

        actions = [
            ("Nouvelle observation",        "➕", VERT,                  "form"),
            ("Liste des espèces",            "📋", VERT2,                 "list"),
            ("Graphiques & Statistiques",    "📊", BLEU,                  "charts"),
            ("Exporter Excel",               "📤", ORANGE,                "excel"),
            ("Sync  Google Sheets",          "☁️",  (0.19,0.66,0.32,1),  "sync"),
            ("Paramètres Synchronisation",   "⚙️",  INDIGO,               "sync_config"),
            ("À propos",                     "ℹ️",  (0.35,0.35,0.35,1),  "about"),
        ]
        for label, icon, color, action in actions:
            btn = styled_btn(f"{icon}  {label}", color=color, height=dp(48))
            btn.bind(on_press=lambda i, a=action: self._action(a))
            btns.add_widget(btn)

        scroll.add_widget(btns)
        root.add_widget(scroll)
        self.add_widget(root)

    def _action(self, action):
        if action == "excel":
            export_excel()
        elif action == "sync":
            sync_google_sheets()
        else:
            self.manager.current = action

# ═══════════════════════════════════════════
# FORMULAIRE — 4 ONGLETS
# ═══════════════════════════════════════════

class SpeciesForm(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.species_id  = None
        self.photos_list = []
        self.current_tab = 0
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation="vertical")

        # Header
        self.header_lbl = Label(text="Nouvelle observation", bold=True,
                                 font_size=dp(17), color=BLANC)
        header = BoxLayout(size_hint_y=None, height=dp(50), padding=[dp(12),dp(6)])
        with header.canvas.before:
            Color(*VERT)
            hr = Rectangle(pos=header.pos, size=header.size)
        header.bind(pos=lambda i,v: setattr(hr,'pos',v),
                    size=lambda i,v: setattr(hr,'size',v))
        header.add_widget(self.header_lbl)
        root.add_widget(header)

        # Onglets
        tab_bar = BoxLayout(size_hint_y=None, height=dp(42))
        self.tab_btns = []
        for i, label in enumerate(["📍 Lieu", "🔬 Taxon", "👁️ Obs.", "📋 Preuves"]):
            btn = Button(text=label, font_size=dp(12), bold=True,
                         background_normal="", color=BLANC,
                         background_color=ONGLET_COLORS[i])
            btn.bind(on_press=lambda x, idx=i: self._show_tab(idx))
            self.tab_btns.append(btn)
            tab_bar.add_widget(btn)
        root.add_widget(tab_bar)

        # Contenu scrollable
        self.scroll = ScrollView()
        self.tabs = []
        for builder in [self._tab_localisation, self._tab_taxonomie,
                        self._tab_observation, self._tab_preuves]:
            c = BoxLayout(orientation="vertical", spacing=dp(6),
                          padding=dp(14), size_hint_y=None)
            c.bind(minimum_height=c.setter("height"))
            builder(c)
            self.tabs.append(c)
        self.scroll.add_widget(self.tabs[0])
        root.add_widget(self.scroll)

        # Navigation
        nav = BoxLayout(size_hint_y=None, height=dp(50),
                        spacing=dp(6), padding=[dp(8),dp(4)])
        bp = styled_btn("◀ Préc.",      color=(0.5,0.5,0.5,1), height=dp(44))
        bs = styled_btn("💾 Enregistrer", color=VERT,           height=dp(44))
        bn = styled_btn("Suiv. ▶",      color=BLEU,            height=dp(44))
        bb = styled_btn("✕",            color=ROUGE,           height=dp(44))
        bb.size_hint_x = None
        bb.width = dp(48)
        bp.bind(on_press=lambda i: self._nav(-1))
        bn.bind(on_press=lambda i: self._nav(+1))
        bs.bind(on_press=self._save)
        bb.bind(on_press=lambda i: setattr(self.manager,"current","menu"))
        nav.add_widget(bb)
        nav.add_widget(bp)
        nav.add_widget(bs)
        nav.add_widget(bn)
        root.add_widget(nav)
        self.add_widget(root)
        self._show_tab(0)

    def _show_tab(self, idx):
        self.current_tab = idx
        self.scroll.clear_widgets()
        self.scroll.scroll_y = 1
        self.scroll.add_widget(self.tabs[idx])
        for i, btn in enumerate(self.tab_btns):
            c = ONGLET_COLORS[i]
            btn.background_color = c if i == idx else (c[0]*.6, c[1]*.6, c[2]*.6, 1)

    def _nav(self, d):
        n = self.current_tab + d
        if 0 <= n <= 3:
            self._show_tab(n)

    # ─── ONGLET 1 : LOCALISATION ────────────
    def _tab_localisation(self, L):
        L.add_widget(section_lbl("Coordonnées GPS"))

        row_gps = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        self.f_lat  = txt("Latitude")
        self.f_lon  = txt("Longitude")
        btn_gps = styled_btn("📍 Auto", color=BLEU, height=dp(44))
        btn_gps.size_hint_x = None
        btn_gps.width = dp(90)
        btn_gps.bind(on_press=self._use_gps)
        row_gps.add_widget(self.f_lat)
        row_gps.add_widget(self.f_lon)
        row_gps.add_widget(btn_gps)
        L.add_widget(field_lbl("Latitude  /  Longitude"))
        L.add_widget(row_gps)

        row_ap = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        self.f_alt  = txt("Altitude (m)")
        self.f_prec = txt("Précision GPS (m)")
        row_ap.add_widget(self.f_alt)
        row_ap.add_widget(self.f_prec)
        L.add_widget(field_lbl("Altitude (m)  /  Précision GPS (m)"))
        L.add_widget(row_ap)

        L.add_widget(section_lbl("Localisation"))
        L.add_widget(field_lbl("Nom du site / Lieu-dit"))
        self.f_lieu = txt("Ex : Mare de Guidimouni")
        L.add_widget(self.f_lieu)

        L.add_widget(field_lbl("Type de milieu (habitat)"))
        self.sp_milieu = make_spinner(MILIEUX)
        L.add_widget(self.sp_milieu)

        L.add_widget(field_lbl("Description du site"))
        self.f_desc = txt("Décris l'environnement…", multiline=True, height=dp(80))
        L.add_widget(self.f_desc)

    # ─── ONGLET 2 : TAXONOMIE ───────────────
    def _tab_taxonomie(self, L):
        L.add_widget(section_lbl("Identification taxonomique"))

        L.add_widget(field_lbl("Règne / Groupe biologique *"))
        self.sp_regne = make_spinner(REGNES)
        L.add_widget(self.sp_regne)

        L.add_widget(field_lbl("Nom scientifique (Latin) *"))
        self.f_nom_sc = txt("Ex : Panthera leo")
        L.add_widget(self.f_nom_sc)

        L.add_widget(field_lbl("Nom vernaculaire / local"))
        self.f_nom_loc = txt("Ex : Lion")
        L.add_widget(self.f_nom_loc)

        L.add_widget(section_lbl("Caractéristiques"))

        L.add_widget(field_lbl("Stade de vie"))
        self.sp_stade = make_spinner(STADES_VIE)
        L.add_widget(self.sp_stade)

        L.add_widget(field_lbl("Sexe"))
        self.sp_sexe = make_spinner(SEXES)
        L.add_widget(self.sp_sexe)

    # ─── ONGLET 3 : OBSERVATION ─────────────
    def _tab_observation(self, L):
        L.add_widget(section_lbl("Horodatage"))

        row_dt = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        self.f_date  = txt("Date (JJ/MM/AAAA)")
        self.f_heure = txt("Heure (HH:MM)")
        btn_now = styled_btn("Maintenant", color=BLEU, height=dp(44))
        btn_now.size_hint_x = None
        btn_now.width = dp(120)
        btn_now.bind(on_press=self._use_datetime)
        row_dt.add_widget(self.f_date)
        row_dt.add_widget(self.f_heure)
        row_dt.add_widget(btn_now)
        L.add_widget(field_lbl("Date  /  Heure"))
        L.add_widget(row_dt)

        L.add_widget(section_lbl("Données d'observation"))

        L.add_widget(field_lbl("Type d'indice"))
        self.sp_indice = make_spinner(TYPES_INDICE)
        L.add_widget(self.sp_indice)

        L.add_widget(field_lbl("Abondance / Classe d'effectif"))
        self.sp_abondance = make_spinner(ABONDANCES)
        L.add_widget(self.sp_abondance)

        L.add_widget(field_lbl("Nombre exact d'individus (optionnel)"))
        self.f_nb_ind = txt("Ex : 3")
        L.add_widget(self.f_nb_ind)

        L.add_widget(field_lbl("Comportement observé"))
        self.sp_comportement = make_spinner(COMPORTEMENTS)
        L.add_widget(self.sp_comportement)

    # ─── ONGLET 4 : PREUVES ─────────────────
    def _tab_preuves(self, L):
        L.add_widget(section_lbl("Photos"))
        self.photo_count_lbl = Label(
            text="Aucune photo", size_hint_y=None, height=dp(22),
            color=(0.5,0.5,0.5,1), font_size=dp(12))
        L.add_widget(self.photo_count_lbl)

        row_ph = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        bg = styled_btn("Galerie",  color=VIOLET, height=dp(44))
        bc = styled_btn("Caméra",   color=(0.3,0.3,0.3,1), height=dp(44))
        be = styled_btn("Effacer",  color=ROUGE,  height=dp(44))
        bg.bind(on_press=self._open_gallery)
        bc.bind(on_press=self._take_photo)
        be.bind(on_press=self._clear_photos)
        row_ph.add_widget(bg)
        row_ph.add_widget(bc)
        row_ph.add_widget(be)
        L.add_widget(row_ph)

        self.thumb_grid = GridLayout(cols=2, spacing=dp(6),
                                      size_hint_y=None, height=dp(0))
        L.add_widget(self.thumb_grid)

        L.add_widget(section_lbl("Audio"))
        L.add_widget(field_lbl("Chemin fichier audio"))
        self.f_audio = txt("Ex : /sdcard/enregistrement_01.mp3")
        L.add_widget(self.f_audio)

        L.add_widget(section_lbl("Métadonnées"))

        L.add_widget(field_lbl("Nom de l'observateur *"))
        self.f_observateur = txt("Ex : Moctar Maman Laouan Nouhou")
        L.add_widget(self.f_observateur)

        L.add_widget(field_lbl("Protocole utilisé"))
        self.sp_protocole = make_spinner(PROTOCOLES)
        L.add_widget(self.sp_protocole)

        L.add_widget(field_lbl("Statut de validation"))
        self.sp_statut = make_spinner(STATUTS_VALIDATION, default="En attente")
        L.add_widget(self.sp_statut)

        L.add_widget(field_lbl("Notes complémentaires"))
        self.f_notes = txt("Toute remarque…", multiline=True, height=dp(80))
        L.add_widget(self.f_notes)

    # ─── GPS / DATETIME ─────────────────────
    def _use_gps(self, *args):
        if gps_data["lat"]:
            self.f_lat.text  = gps_data["lat"]
            self.f_lon.text  = gps_data["lon"]
            self.f_alt.text  = gps_data["alt"]
            self.f_prec.text = gps_data["precision"]
        else:
            show_popup("GPS", "Coordonnées pas encore disponibles.", ORANGE)

    def _use_datetime(self, *args):
        now = datetime.now()
        self.f_date.text  = now.strftime("%d/%m/%Y")
        self.f_heure.text = now.strftime("%H:%M")

    # ─── GALERIE ────────────────────────────
    def _open_gallery(self, *args):
        start = "/sdcard/DCIM" if os.path.exists("/sdcard/DCIM") else os.path.expanduser("~")
        content = BoxLayout(orientation="vertical", spacing=dp(6), padding=dp(8))
        fc = FileChooserIconView(path=start,
             filters=["*.jpg","*.jpeg","*.png","*.JPG","*.JPEG","*.PNG"])
        row = BoxLayout(size_hint_y=None, height=dp(46), spacing=dp(8))
        bok = styled_btn("Sélectionner", color=VERT)
        ban = styled_btn("Annuler",      color=ROUGE)
        row.add_widget(bok)
        row.add_widget(ban)
        content.add_widget(fc)
        content.add_widget(row)
        popup = Popup(title="Sélectionner une photo", content=content,
                      size_hint=(0.95,0.9), background="",
                      background_color=(0.1,0.1,0.1,0.95), title_color=BLANC)
        ban.bind(on_press=popup.dismiss)
        def _sel(*a):
            if fc.selection: self._add_photo(fc.selection[0])
            popup.dismiss()
        bok.bind(on_press=_sel)
        popup.open()

    def _take_photo(self, *args):
        if not PLYER_OK:
            show_popup("Caméra", "plyer non disponible.", ORANGE)
            return
        fn = os.path.join(os.path.dirname(os.path.abspath(__file__)), "photo_obs.jpg")
        try:
            camera.take_picture(fn, lambda p: self._add_photo(p))
        except Exception as e:
            show_popup("Caméra", str(e), ROUGE)

    def _add_photo(self, path):
        if path and path not in self.photos_list:
            self.photos_list.append(path)
        self._refresh_thumbs()

    def _clear_photos(self, *args):
        self.photos_list = []
        self._refresh_thumbs()

    def _remove_photo(self, path):
        if path in self.photos_list: self.photos_list.remove(path)
        self._refresh_thumbs()

    def _refresh_thumbs(self):
        self.thumb_grid.clear_widgets()
        n = len(self.photos_list)
        self.photo_count_lbl.text = f"{n} photo(s)" if n else "Aucune photo"
        rows = (n+1)//2 if n else 0
        self.thumb_grid.height = dp(rows*130 + max(0,rows-1)*6)
        for path in self.photos_list:
            box = BoxLayout(orientation="vertical", size_hint_y=None,
                            height=dp(128), spacing=dp(2))
            box.add_widget(Image(source=path, allow_stretch=True, keep_ratio=True,
                                  size_hint_y=None, height=dp(100)))
            box.add_widget(Label(text=os.path.basename(path)[:20], font_size=dp(9),
                                  color=(0.4,0.4,0.4,1), size_hint_y=None, height=dp(14)))
            br = styled_btn("✕", color=ROUGE, height=dp(22))
            br.font_size = dp(10)
            br.bind(on_press=lambda i, p=path: self._remove_photo(p))
            box.add_widget(br)
            self.thumb_grid.add_widget(box)

    # ─── CHARGEMENT MODE ÉDITION ────────────
    def load_species(self, sid):
        self.species_id = sid
        self.header_lbl.text = "Modifier l'observation"
        cursor.execute("SELECT * FROM species WHERE id=?", (sid,))
        row = cursor.fetchone()
        if not row: return

        def g(i, d=""):
            return row[i] if len(row) > i and row[i] else d

        self.f_lat.text           = g(5)
        self.f_lon.text           = g(6)
        self.f_alt.text           = g(9)
        self.f_prec.text          = g(10)
        self.f_lieu.text          = g(4)
        self.sp_milieu.text       = g(11) or MILIEUX[0]
        self.f_desc.text          = g(7)
        self.sp_regne.text        = g(3)  or REGNES[0]
        self.f_nom_sc.text        = g(1)
        self.f_nom_loc.text       = g(2)
        self.sp_stade.text        = g(12) or STADES_VIE[0]
        self.sp_sexe.text         = g(13) or SEXES[0]
        self.f_date.text          = g(14)
        self.f_heure.text         = g(15)
        self.sp_indice.text       = g(16) or TYPES_INDICE[0]
        self.sp_abondance.text    = g(17) or ABONDANCES[0]
        self.f_nb_ind.text        = g(19)
        self.sp_comportement.text = g(18) or COMPORTEMENTS[0]
        self.f_audio.text         = g(20)
        self.f_observateur.text   = g(21)
        self.sp_protocole.text    = g(22) or PROTOCOLES[0]
        self.sp_statut.text       = g(23) or "En attente"
        self.f_notes.text         = g(24)
        self.photos_list = [p.strip() for p in g(8).split("|") if p.strip()]
        self._refresh_thumbs()
        self._show_tab(0)

    def reset_form(self):
        self.species_id = None
        self.header_lbl.text = "Nouvelle observation"
        for f in [self.f_lat, self.f_lon, self.f_alt, self.f_prec,
                  self.f_lieu, self.f_desc, self.f_nom_sc, self.f_nom_loc,
                  self.f_date, self.f_heure, self.f_nb_ind,
                  self.f_audio, self.f_observateur, self.f_notes]:
            f.text = ""
        for sp, lst in [(self.sp_milieu, MILIEUX), (self.sp_regne, REGNES),
                         (self.sp_stade, STADES_VIE), (self.sp_sexe, SEXES),
                         (self.sp_indice, TYPES_INDICE), (self.sp_abondance, ABONDANCES),
                         (self.sp_comportement, COMPORTEMENTS), (self.sp_protocole, PROTOCOLES)]:
            sp.text = lst[0]
        self.sp_statut.text = "En attente"
        self.photos_list = []
        self._refresh_thumbs()
        self._show_tab(0)

    # ─── ENREGISTREMENT ─────────────────────
    def _save(self, *args):
        if not self.f_nom_sc.text.strip():
            show_popup("Champ requis",
                       "Nom scientifique obligatoire.\n(Onglet Taxon)", ROUGE)
            self._show_tab(1)
            return
        if not self.f_observateur.text.strip():
            show_popup("Champ requis",
                       "Nom de l'observateur obligatoire.\n(Onglet Preuves)", ROUGE)
            self._show_tab(3)
            return

        def sv(sp, lst):
            v = sp.text
            return "" if v == lst[0] else v

        # Ordre d'insertion :
        # 1=nom_sc, 2=nom_loc, 3=type(règne), 4=lieu,
        # 5=lat, 6=lon, 7=desc, 8=photos,
        # 9=altitude, 10=precision_gps, 11=type_milieu,
        # 12=stade_vie, 13=sexe,
        # 14=date_obs, 15=heure_obs,
        # 16=type_indice, 17=abondance, 18=comportement, 19=nb_individus,
        # 20=audio, 21=observateur, 22=protocole, 23=statut_validation, 24=notes

        vals = (
            self.f_nom_sc.text.strip(),
            self.f_nom_loc.text.strip(),
            sv(self.sp_regne, REGNES),
            self.f_lieu.text.strip(),
            self.f_lat.text.strip(),
            self.f_lon.text.strip(),
            self.f_desc.text.strip(),
            "|".join(self.photos_list),
            self.f_alt.text.strip(),
            self.f_prec.text.strip(),
            sv(self.sp_milieu, MILIEUX),
            sv(self.sp_stade, STADES_VIE),
            sv(self.sp_sexe, SEXES),
            self.f_date.text.strip(),
            self.f_heure.text.strip(),
            sv(self.sp_indice, TYPES_INDICE),
            sv(self.sp_abondance, ABONDANCES),
            sv(self.sp_comportement, COMPORTEMENTS),
            self.f_nb_ind.text.strip(),
            self.f_audio.text.strip(),
            self.f_observateur.text.strip(),
            sv(self.sp_protocole, PROTOCOLES),
            self.sp_statut.text.strip(),
            self.f_notes.text.strip(),
        )

        try:
            if self.species_id:
                cursor.execute("""
                    UPDATE species SET
                        nom_scientifique=?, nom_local=?, type=?, lieu=?,
                        latitude=?, longitude=?, description=?, photos=?,
                        altitude=?, precision_gps=?, type_milieu=?,
                        stade_vie=?, sexe=?,
                        date_obs=?, heure_obs=?,
                        type_indice=?, abondance=?, comportement=?, nb_individus=?,
                        audio=?, observateur=?, protocole=?,
                        statut_validation=?, notes=?
                    WHERE id=?
                """, vals + (self.species_id,))
                msg = "Observation modifiée."
            else:
                cursor.execute("""
                    INSERT INTO species (
                        nom_scientifique, nom_local, type, lieu,
                        latitude, longitude, description, photos,
                        altitude, precision_gps, type_milieu,
                        stade_vie, sexe,
                        date_obs, heure_obs,
                        type_indice, abondance, comportement, nb_individus,
                        audio, observateur, protocole,
                        statut_validation, notes
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, vals)
                msg = "Observation enregistrée !"
            conn.commit()
            show_popup("Succès", msg, VERT)
            self.reset_form()
            self.manager.current = "menu"
        except Exception as e:
            show_popup("Erreur DB", str(e), ROUGE)

# ═══════════════════════════════════════════
# ÉCRAN : LISTE
# ═══════════════════════════════════════════

class ListScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.all_data = []
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation="vertical")
        root.add_widget(make_header("📋 Liste des observations", VERT))

        search_row = BoxLayout(size_hint_y=None, height=dp(44),
                               padding=[dp(6),dp(2)], spacing=dp(6))
        self.search_input = TextInput(hint_text="🔍 Rechercher…",
                                       multiline=False, font_size=dp(13))
        self.search_input.bind(text=lambda i,v: self._filter())
        self.sp_filtre = make_spinner(
            ["Tous","En attente","Validé","Rejeté","À vérifier"], default="Tous")
        self.sp_filtre.size_hint_x = None
        self.sp_filtre.width = dp(120)
        self.sp_filtre.bind(text=lambda i,v: self._filter())
        search_row.add_widget(self.search_input)
        search_row.add_widget(self.sp_filtre)
        root.add_widget(search_row)

        self.scroll = ScrollView()
        self.list_layout = BoxLayout(orientation="vertical", spacing=dp(6),
                                      padding=[dp(6),dp(4)], size_hint_y=None)
        self.list_layout.bind(minimum_height=self.list_layout.setter("height"))
        self.scroll.add_widget(self.list_layout)
        root.add_widget(self.scroll)

        btn_row = BoxLayout(size_hint_y=None, height=dp(50),
                            spacing=dp(8), padding=[dp(10),dp(4)])
        br = styled_btn("🔄 Actualiser", color=BLEU)
        bb = styled_btn("← Retour",      color=(0.5,0.5,0.5,1))
        br.bind(on_press=lambda i: self.load())
        bb.bind(on_press=lambda i: setattr(self.manager,"current","menu"))
        btn_row.add_widget(br)
        btn_row.add_widget(bb)
        root.add_widget(btn_row)
        self.add_widget(root)

    def on_enter(self):
        self.load()

    def load(self):
        cursor.execute("SELECT * FROM species ORDER BY id DESC")
        self.all_data = cursor.fetchall()
        self._filter()

    def _filter(self):
        q  = self.search_input.text.lower().strip()
        st = self.sp_filtre.text

        def match(row):
            def g(i): return (row[i] or "").lower() if len(row) > i else ""
            text_ok = not q or any(q in g(i) for i in [1,2,3,4,21])
            stat_ok = st == "Tous" or g(23).lower() == st.lower()
            return text_ok and stat_ok

        self._render([r for r in self.all_data if match(r)])

    def _render(self, rows):
        self.list_layout.clear_widgets()
        if not rows:
            self.list_layout.add_widget(Label(
                text="Aucune observation trouvée.",
                size_hint_y=None, height=dp(60), color=(0.5,0.5,0.5,1)))
            return
        for row in rows:
            self.list_layout.add_widget(self._make_card(row))

    def _make_card(self, row):
        def g(i, d=""): return row[i] if len(row) > i and row[i] else d

        photos = [p.strip() for p in g(8).split("|") if p.strip()]
        statut = g(23, "En attente")
        s_color = {"Validé":(0.13,0.55,0.13,1),"Rejeté":ROUGE,
                    "À vérifier":ORANGE}.get(statut, (0.5,0.5,0.5,1))
        has_photos = bool(photos)
        card_h = dp(150) + (dp(108) if has_photos else 0)

        card = BoxLayout(orientation="vertical", size_hint_y=None,
                         height=card_h, padding=dp(10), spacing=dp(4))
        with card.canvas.before:
            Color(*GRIS_FOND)
            cr = Rectangle(pos=card.pos, size=card.size)
        card.bind(pos=lambda i,v: setattr(cr,'pos',v),
                  size=lambda i,v: setattr(cr,'size',v))

        # Ligne infos + miniature
        top = BoxLayout(size_hint_y=None, height=dp(100), spacing=dp(8))
        info = BoxLayout(orientation="vertical", spacing=dp(2))

        def sl(text, color=(0.15,0.15,0.15,1), bold=False, fs=dp(13)):
            l = Label(text=text, font_size=fs, color=color, bold=bold,
                      halign="left", size_hint_y=None, height=dp(18))
            l.bind(size=lambda i,v: setattr(l,'text_size',v))
            return l

        info.add_widget(sl(f"{g(1)}  —  {g(2)}", bold=True, fs=dp(14), color=(0.1,0.1,0.1,1)))
        info.add_widget(sl(f"🦁 {g(3)}   📍 {g(4)}", color=(0.3,0.3,0.3,1)))
        info.add_widget(sl(f"📅 {g(14)} {g(15)}   👤 {g(21)}", color=(0.4,0.4,0.4,1), fs=dp(11)))
        info.add_widget(sl(f"🌿 {g(11)}   📊 {g(17)}", color=(0.4,0.4,0.4,1), fs=dp(11)))
        info.add_widget(sl(f"GPS: {g(5)}/{g(6)}  Alt:{g(9)}m  Prec:{g(10)}m",
                            color=(0.45,0.45,0.45,1), fs=dp(10)))
        badge = Label(text=statut, font_size=dp(10), bold=True, color=s_color,
                      size_hint_y=None, height=dp(16), halign="left")
        badge.bind(size=lambda i,v: setattr(badge,'text_size',v))
        info.add_widget(badge)
        top.add_widget(info)

        if photos and os.path.exists(photos[0]):
            top.add_widget(Image(source=photos[0], allow_stretch=True,
                                  keep_ratio=True, size_hint_x=None, width=dp(80)))
        card.add_widget(top)

        if has_photos:
            sh = ScrollView(size_hint_y=None, height=dp(100),
                            do_scroll_x=True, do_scroll_y=False)
            gallery = BoxLayout(orientation="horizontal", spacing=dp(6),
                                size_hint_x=None, width=dp(len(photos)*106))
            for p in photos:
                if os.path.exists(p):
                    ib = BoxLayout(orientation="vertical",
                                   size_hint_x=None, width=dp(100), spacing=dp(2))
                    ib.add_widget(Image(source=p, allow_stretch=True, keep_ratio=True,
                                        size_hint_y=None, height=dp(80)))
                    ib.add_widget(Label(text=os.path.basename(p)[:14], font_size=dp(9),
                                        color=(0.5,0.5,0.5,1), size_hint_y=None, height=dp(14)))
                    gallery.add_widget(ib)
            sh.add_widget(gallery)
            card.add_widget(sh)

        btns = BoxLayout(size_hint_y=None, height=dp(36), spacing=dp(6))
        be = styled_btn("✏️ Modifier",  color=VERT,  height=dp(36))
        bd = styled_btn("🗑️ Supprimer", color=ROUGE, height=dp(36))
        be.bind(on_press=lambda i, rid=row[0]: self._edit(rid))
        bd.bind(on_press=lambda i, rid=row[0]: self._confirm_delete(rid))
        btns.add_widget(be)
        btns.add_widget(bd)
        card.add_widget(btns)
        return card

    def _edit(self, sid):
        self.manager.get_screen("form").load_species(sid)
        self.manager.current = "form"

    def _confirm_delete(self, sid):
        content = BoxLayout(orientation="vertical", padding=dp(16), spacing=dp(10))
        content.add_widget(Label(text="Supprimer cette observation ?",
                                  color=(0.9,0.9,0.9,1)))
        row = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        by = styled_btn("Oui, supprimer", color=ROUGE)
        bn = styled_btn("Annuler",        color=(0.5,0.5,0.5,1))
        row.add_widget(by)
        row.add_widget(bn)
        content.add_widget(row)
        popup = Popup(title="Confirmation", content=content,
                      size_hint=(0.82,None), height=dp(200),
                      background="", background_color=(0.1,0.1,0.1,0.95),
                      title_color=BLANC)
        bn.bind(on_press=popup.dismiss)
        def _do(*a):
            cursor.execute("DELETE FROM species WHERE id=?", (sid,))
            conn.commit()
            popup.dismiss()
            self.load()
        by.bind(on_press=_do)
        popup.open()

# ═══════════════════════════════════════════
# ÉCRAN : GRAPHIQUES
# ═══════════════════════════════════════════

class ChartsScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation="vertical")
        root.add_widget(make_header("📊 Graphiques & Statistiques", BLEU))
        self.scroll = ScrollView()
        self.inner  = BoxLayout(orientation="vertical", spacing=dp(16),
                                 padding=dp(12), size_hint_y=None)
        self.inner.bind(minimum_height=self.inner.setter("height"))
        self.scroll.add_widget(self.inner)
        root.add_widget(self.scroll)
        btn_row = BoxLayout(size_hint_y=None, height=dp(50),
                            spacing=dp(8), padding=[dp(10),dp(4)])
        br = styled_btn("🔄 Actualiser", color=BLEU)
        bb = styled_btn("← Retour",      color=(0.5,0.5,0.5,1))
        br.bind(on_press=lambda i: self.load())
        bb.bind(on_press=lambda i: setattr(self.manager,"current","menu"))
        btn_row.add_widget(br)
        btn_row.add_widget(bb)
        root.add_widget(btn_row)
        self.add_widget(root)

    def on_enter(self):
        self.load()

    def load(self):
        self.inner.clear_widgets()
        cursor.execute("SELECT type, lieu, type_milieu, statut_validation FROM species")
        rows = cursor.fetchall()
        if not rows:
            self.inner.add_widget(Label(text="Aucune donnée.",
                                        color=(0.5,0.5,0.5,1),
                                        size_hint_y=None, height=dp(60)))
            return

        def agg(idx):
            d = {}
            for r in rows:
                k = (r[idx] or "Inconnu").strip() or "Inconnu"
                d[k] = d.get(k,0) + 1
            return d

        # Bandeau stats
        nb = len(rows)
        sb = BoxLayout(size_hint_y=None, height=dp(56), padding=dp(10))
        with sb.canvas.before:
            Color(*VERT)
            sr = Rectangle(pos=sb.pos, size=sb.size)
        sb.bind(pos=lambda i,v: setattr(sr,'pos',v),
                size=lambda i,v: setattr(sr,'size',v))
        sb.add_widget(Label(
            text=f"Total : {nb} observation(s)   |   "
                 f"{len(agg(0))} règne(s)   |   {len(agg(1))} site(s)",
            color=BLANC, bold=True, font_size=dp(13)))
        self.inner.add_widget(sb)

        for data, title in [
            (agg(0), "Répartition par Règne/Groupe"),
            (agg(1), "Observations par Site"),
            (agg(2), "Répartition par Milieu"),
            (agg(3), "Statuts de validation"),
        ]:
            if len(data) <= 6:
                w = PieChartWidget(data, title=title, size_hint_y=None, height=dp(280))
            else:
                w = BarChartWidget(data, title=title, size_hint_y=None, height=dp(240))
            self.inner.add_widget(w)

# ═══════════════════════════════════════════
# ÉCRAN : PARAMÈTRES SYNC
# ═══════════════════════════════════════════

class SyncConfigScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation="vertical")
        root.add_widget(make_header("⚙️ Paramètres Synchronisation", INDIGO))
        scroll = ScrollView()
        content = BoxLayout(orientation="vertical", spacing=dp(14),
                             padding=dp(16), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))

        for num, titre, texte in [
            ("1️⃣","Créer un Google Sheet",
             "sheets.google.com → Nouveau tableau\n"
             "Copie l'ID dans l'URL (entre /d/ et /edit)"),
            ("2️⃣","Créer le script Apps Script",
             "script.google.com → Nouveau projet\n"
             "Colle le code ci-dessous, remplace TON_ID"),
            ("3️⃣","Déployer",
             "Déployer → Application Web\n"
             "Accès : Tout le monde → Copie l'URL"),
            ("4️⃣","Coller l'URL ci-dessous",
             "Entre l'URL et clique Enregistrer"),
        ]:
            card = BoxLayout(orientation="vertical", size_hint_y=None,
                              height=dp(80+texte.count('\n')*18),
                              padding=dp(10), spacing=dp(4))
            with card.canvas.before:
                Color(0.93,0.93,1.0,1)
                cr = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=lambda i,v,cr=cr: setattr(cr,'pos',v),
                      size=lambda i,v,cr=cr: setattr(cr,'size',v))
            row = BoxLayout(size_hint_y=None, height=dp(24), spacing=dp(6))
            row.add_widget(Label(text=num, font_size=dp(18),
                                  size_hint_x=None, width=dp(32)))
            row.add_widget(Label(text=titre, bold=True, font_size=dp(13),
                                  color=INDIGO, halign="left"))
            lbl = Label(text=texte, font_size=dp(12),
                        color=(0.2,0.2,0.2,1), halign="left", size_hint_y=None)
            lbl.bind(texture_size=lbl.setter("size"))
            card.add_widget(row)
            card.add_widget(lbl)
            content.add_widget(card)

        content.add_widget(field_lbl("🔗 URL Google Apps Script :"))
        self.url_input = TextInput(
            hint_text="https://script.google.com/macros/s/…",
            text="" if GOOGLE_SHEET_URL == "COLLE_ICI_TON_URL_GOOGLE_APPS_SCRIPT"
                  else GOOGLE_SHEET_URL,
            multiline=False, size_hint_y=None, height=dp(46), font_size=dp(11))
        content.add_widget(self.url_input)

        bt = styled_btn("🧪 Tester la connexion", color=BLEU)
        bt.size_hint_y = None
        bt.height = dp(46)
        bt.bind(on_press=self._test)
        content.add_widget(bt)

        content.add_widget(section_lbl("Code Apps Script à copier :"))
        code_box = BoxLayout(size_hint_y=None, height=dp(400), padding=dp(8))
        with code_box.canvas.before:
            Color(0.08,0.08,0.08,1)
            cbr = Rectangle(pos=code_box.pos, size=code_box.size)
        code_box.bind(pos=lambda i,v: setattr(cbr,'pos',v),
                      size=lambda i,v: setattr(cbr,'size',v))
        code = (
            "function doPost(e) {\n"
            "  var sheet = SpreadsheetApp\n"
            "    .openById('TON_ID_SHEET')\n"
            "    .getActiveSheet();\n"
            "  var d = JSON.parse(\n"
            "    e.postData.contents);\n"
            "  sheet.appendRow([\n"
            "    new Date(),\n"
            "    d.nom_scientifique,\n"
            "    d.nom_local,\n"
            "    d.type,\n"
            "    d.lieu,\n"
            "    d.latitude,\n"
            "    d.longitude,\n"
            "    d.altitude,\n"
            "    d.precision_gps,\n"
            "    d.type_milieu,\n"
            "    d.date_obs,\n"
            "    d.heure_obs,\n"
            "    d.type_indice,\n"
            "    d.abondance,\n"
            "    d.nb_individus,\n"
            "    d.comportement,\n"
            "    d.observateur,\n"
            "    d.protocole,\n"
            "    d.statut_validation,\n"
            "    d.nb_photos,\n"
            "    d.notes\n"
            "  ]);\n"
            "  return ContentService\n"
            "    .createTextOutput(\n"
            "      JSON.stringify(\n"
            "        {status:'ok'}))\n"
            "    .setMimeType(\n"
            "      ContentService\n"
            "      .MimeType.JSON);\n"
            "}"
        )
        cl = Label(text=code, color=(0.2,1,0.4,1), font_size=dp(10),
                    halign="left", valign="top")
        cl.bind(size=lambda i,v: setattr(cl,'text_size',v))
        code_box.add_widget(cl)
        content.add_widget(code_box)

        scroll.add_widget(content)
        root.add_widget(scroll)

        btn_row = BoxLayout(size_hint_y=None, height=dp(50),
                            spacing=dp(8), padding=[dp(10),dp(4)])
        bs = styled_btn("💾 Enregistrer", color=VERT)
        bb = styled_btn("← Retour",       color=(0.5,0.5,0.5,1))
        bs.bind(on_press=self._save_url)
        bb.bind(on_press=lambda i: setattr(self.manager,"current","menu"))
        btn_row.add_widget(bs)
        btn_row.add_widget(bb)
        root.add_widget(btn_row)
        self.add_widget(root)

    def _save_url(self, *args):
        global GOOGLE_SHEET_URL
        url = self.url_input.text.strip()
        if not url.startswith("https://"):
            show_popup("URL invalide", "L'URL doit commencer par https://", ROUGE)
            return
        GOOGLE_SHEET_URL = url
        show_popup("Enregistré", "URL sauvegardée !", VERT)

    def _test(self, *args):
        url = self.url_input.text.strip()
        if not url.startswith("https://"):
            show_popup("Erreur", "Entre d'abord une URL valide.", ROUGE)
            return
        try:
            r = requests.post(url, json={"nom_scientifique":"TEST",
                              "observateur":"Test","nb_photos":0},
                              timeout=10, allow_redirects=True)
            if r.status_code == 200:
                show_popup("Connexion OK",
                           "Serveur OK ! Une ligne TEST\na été ajoutée.", VERT)
            else:
                show_popup("Réponse inattendue",
                           f"Code HTTP : {r.status_code}", ORANGE)
        except Exception as e:
            show_popup("Erreur réseau", str(e)[:100], ROUGE)

# ═══════════════════════════════════════════
# ÉCRAN : À PROPOS
# ═══════════════════════════════════════════

class AboutScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        root = BoxLayout(orientation="vertical")
        root.add_widget(make_header("ℹ️  À propos", INDIGO))
        scroll = ScrollView()
        content = BoxLayout(orientation="vertical", spacing=dp(14),
                             padding=dp(20), size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))
        content.add_widget(Label(text="🎓", font_size=dp(56),
                                  size_hint_y=None, height=dp(70)))

        for titre, texte in [
            ("🏛️  Institution",
             "Université André Salifou de Zinder\nZinder, Niger"),
            ("🌿  Application",
             "Inventaire de la Biodiversité\nVersion 3.0"),
            ("👤  Développeur",
             "Moctar Maman Laouan Nouhou\nMoctar3490@gmail.com"),
            ("🛠️  Technologies",
             "Python 3  •  Kivy  •  SQLite\nopenpyxl  •  plyer  •  requests"),
            ("📋  Fonctionnalités",
             "• 4 onglets : Localisation, Taxonomie, Observation, Preuves\n"
             "• GPS auto (lat, lon, altitude, précision)\n"
             "• Date/heure automatique\n"
             "• Multi-photos + galerie + caméra\n"
             "• Listes déroulantes standardisées\n"
             "• Statut de validation\n"
             "• Export Excel professionnel\n"
             "• Synchronisation Google Sheets\n"
             "• 4 graphiques statistiques\n"
             "• Recherche & filtre\n"
             "• 100% offline"),
            ("📅  Année", "2025 — 2026"),
        ]:
            h = dp(80 + texte.count('\n')*18)
            card = BoxLayout(orientation="vertical", size_hint_y=None,
                              height=h, padding=dp(12), spacing=dp(4))
            with card.canvas.before:
                Color(0.94,0.94,1.0,1)
                cr = Rectangle(pos=card.pos, size=card.size)
            card.bind(pos=lambda i,v,cr=cr: setattr(cr,'pos',v),
                      size=lambda i,v,cr=cr: setattr(cr,'size',v))
            card.add_widget(Label(text=titre, bold=True, font_size=dp(13),
                                   color=INDIGO, halign="left",
                                   size_hint_y=None, height=dp(22)))
            lbl = Label(text=texte, font_size=dp(13), color=(0.15,0.15,0.15,1),
                        halign="left", size_hint_y=None)
            lbl.bind(texture_size=lbl.setter("size"))
            card.add_widget(lbl)
            content.add_widget(card)

        content.add_widget(Label(
            text="© 2026 — Université André Salifou de Zinder\nTous droits réservés",
            font_size=dp(11), color=(0.5,0.5,0.5,1),
            halign="center", size_hint_y=None, height=dp(40)))
        scroll.add_widget(content)
        root.add_widget(scroll)
        bb = styled_btn("← Retour", color=INDIGO)
        bb.size_hint_y = None
        bb.height = dp(50)
        bb.bind(on_press=lambda i: setattr(self.manager,"current","menu"))
        root.add_widget(bb)
        self.add_widget(root)

# ═══════════════════════════════════════════
# APPLICATION PRINCIPALE
# ═══════════════════════════════════════════

class InventaireApp(App):
    def build(self):
        db_path = init_db()
        print(f"Base de données : {db_path}")
        start_gps()
        sm = ScreenManager()
        sm.add_widget(MenuScreen(name="menu"))
        sm.add_widget(SpeciesForm(name="form"))
        sm.add_widget(ListScreen(name="list"))
        sm.add_widget(ChartsScreen(name="charts"))
        sm.add_widget(SyncConfigScreen(name="sync_config"))
        sm.add_widget(AboutScreen(name="about"))
        return sm

    def on_stop(self):
        if conn:
            conn.close()
        if PLYER_OK:
            try:
                gps.stop()
            except Exception:
                pass

if __name__ == "__main__":
    InventaireApp().run()
